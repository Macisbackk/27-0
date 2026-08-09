#!/usr/bin/env python3
"""Injury / games-missed aware recalibration of Current peak ratings.

Fantasy column M = metres (NOT matches). Availability is inferred from:
  - Last Rd = 0 + low season Score vs Fantasy price band
  - Very low metres for price
  - xlsx rationale / confidence / form evidence mentioning injury / missed games

Injured / limited players are not crushed for low season totals.
Form-based cuts from the prior audit stay when the player looks available.

Preserves explicit manual locks (Sutton / Loghan / Ashall-Bott) and any current
peak that already differs from a proposed target (treated as manual).

Usage:
  python scripts/calibrate-fantasy-ratings-injury-aware-2026.py
  python scripts/calibrate-fantasy-ratings-injury-aware-2026.py --write-batch
"""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FANTASY_MD = Path(
    r"C:\Users\Macis\.cursor\projects\c-Users-Macis-Projects-27-0\agent-tools\a69fd1e7-4d18-41fb-9e73-7305eee4c674.txt"
)
OUT_BATCH = ROOT / "data" / "player-attr-batch.json"
OUT_REPORT = ROOT / "data" / "fantasy-rating-injury-aware-2026-report.json"
OUT_PRESERVE = ROOT / "data" / "player-rating-manual-preserve-snapshot-2026.json"
PRIOR_REPORT = ROOT / "data" / "fantasy-rating-calibration-2026-report.json"

POS = sorted(
    [
        "Full Back",
        "Winger",
        "Centre",
        "Stand Off",
        "Scrum Half",
        "Prop",
        "Hooker",
        "Second Row",
        "Loose Forward",
    ],
    key=len,
    reverse=True,
)
POS_PAT = "|".join(re.escape(p) for p in POS)

# Explicit user / intentional locks — never overwrite.
MANUAL_LOCKS: dict[str, int] = {
    "bradford-cur-ryan-sutton": 84,
    "bradford-cur-loghan-lewis": 78,
    "toulouse-cur-olly-ashall-bott": 87,
}

# Pre-first-audit peaks for players we wrongly crushed on season totals.
PRE_AUDIT_PEAK: dict[str, int] = {
    "huddersfield-cur-matty-english": 86,
    "huddersfield-cur-niall-evalds": 81,
    "wigan-cur-liam-farrell": 85,
    "hull-fc-cur-herman-eseese": 89,
    "wigan-cur-jai-field": 86,
}

INJURY_RE = re.compile(
    r"injur|surgery|rehab|sidelined|unavailable|hamstring|fracture|"
    r"reduced availability|missed (most|much|games|rounds|time)|"
    r"\bankle\b|\bknee\b|\bACL\b",
    re.I,
)


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", s)).strip()


def parse_price(raw: str) -> int | None:
    t = (raw or "").strip().upper().replace("£", "")
    m = re.match(r"([\d.]+)\s*([KM])?", t)
    if not m:
        return None
    n = float(m.group(1))
    if m.group(2) == "K":
        n *= 1_000
    elif m.group(2) == "M":
        n *= 1_000_000
    return int(round(n))


def expected_score_floor(price: int | None) -> int:
    if price is None:
        return 200
    if price >= 150_000:
        return 700
    if price >= 130_000:
        return 550
    if price >= 110_000:
        return 400
    if price >= 90_000:
        return 250
    return 100


def parse_fantasy(text: str) -> list[dict]:
    """Parse markdown Fantasy table. M = metres, not matches."""
    out: list[dict] = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        parts = [p.strip() for p in line.strip().strip("|").split("|")]
        if len(parts) < 10:
            continue
        # Expected: Team | Name+Pos | Pos# | Price | M | C | TK | ... | Last Rd | Score
        name_cell = parts[1]
        if name_cell in ("Name", "---") or parts[0] == "Team":
            continue
        m = re.match(rf"^(.+?)({POS_PAT})$", name_cell)
        name = m.group(1).strip() if m else name_cell
        try:
            last_rd = int(parts[-2].replace(",", ""))
            score = int(parts[-1].replace(",", ""))
        except ValueError:
            continue
        try:
            metres = int(parts[4].replace(",", ""))
            carries = int(parts[5].replace(",", ""))
            tackles = int(parts[6].replace(",", ""))
        except (ValueError, IndexError):
            metres = carries = tackles = 0
        price_raw = parts[3] if len(parts) > 3 else ""
        out.append(
            {
                "name": name,
                "norm": norm(name),
                "priceRaw": price_raw,
                "price": parse_price(price_raw),
                "metres": metres,
                "carries": carries,
                "tackles": tackles,
                "lastRd": last_rd,
                "score": score,
            }
        )
    return out


def load_xlsx() -> dict[str, dict]:
    wb = openpyxl.load_workbook(ROOT / "super_league_2026_ratings.xlsx", data_only=True)
    ws = wb["All Players"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    xlsx: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("Player"):
            xlsx[norm(str(d["Player"]))] = d
    return xlsx


def xlsx_injury(row: dict | None) -> bool:
    if not row:
        return False
    blob = " ".join(
        str(row.get(k) or "")
        for k in (
            "Rating Rationale",
            "Confidence",
            "Form Evidence",
            "Squad Status",
            "Why",
        )
    )
    return bool(INJURY_RE.search(blob))


def classify(live: dict | None, xrow: dict | None) -> tuple[str, list[str]]:
    notes: list[str] = []
    if not live:
        return "unknown", ["No live Fantasy row"]

    price = live.get("price")
    score = int(live.get("score") or 0)
    last_rd = live.get("lastRd")
    metres = int(live.get("metres") or 0)
    floor = expected_score_floor(price)
    inj = xlsx_injury(xrow)
    if inj:
        notes.append("xlsx mentions injury / reduced availability / missed games")

    if last_rd == 0 and score < floor * 0.55 and (price or 0) >= 90_000:
        notes.append(f"Last Rd=0 and Score {score} << ~{floor} for {live.get('priceRaw')}")
        return "injured_out", notes

    if last_rd == 0 and (price or 0) >= 110_000 and score < floor:
        notes.append(f"Last Rd=0 at {live.get('priceRaw')}; Score {score} below ~{floor}")
        return ("injured_out" if inj or score < floor * 0.75 else "limited"), notes

    if metres < 40 and (price or 0) >= 110_000 and score < floor * 0.7:
        notes.append(f"Very low metres ({metres}) + weak Score {score}")
        return "injured_out", notes

    if inj and score < floor * 0.85:
        notes.append("xlsx injury + below-par season total → limited")
        return "limited", notes

    if last_rd == 0 and (price or 0) >= 100_000:
        notes.append("Missed most recent round (season totals still competitive)")
        return "limited", notes

    notes.append("Looks available / contributing recently")
    return "available", notes


def load_current_players() -> list[dict]:
    raw = json.loads((ROOT / "data" / "current-squads.json").read_text(encoding="utf-8"))
    # Support both {clubs:[{players}]} and flat list shapes
    players: list[dict] = []
    if isinstance(raw, list):
        for p in raw:
            if "-cur-" in str(p.get("id", "")):
                players.append(
                    {
                        "id": p["id"],
                        "name": p["name"],
                        "club": p.get("club") or p.get("team") or "",
                        "peakRating": int(p["peakRating"]),
                    }
                )
        return players
    for club in raw.get("clubs", []):
        club_name = club.get("name") or ""
        for p in club.get("players", []):
            if "-cur-" not in str(p.get("id", "")):
                continue
            players.append(
                {
                    "id": p["id"],
                    "name": p["name"],
                    "club": club_name,
                    "peakRating": int(p["peakRating"]),
                }
            )
    return players


def find_live(live_by: dict[str, dict], name: str) -> dict | None:
    nk = norm(name)
    if nk in live_by:
        return live_by[nk]
    for k, v in live_by.items():
        if nk in k or k in nk:
            return v
    return None


def propose(
    player: dict,
    live: dict | None,
    xrow: dict | None,
    availability: str,
) -> tuple[int | None, str]:
    pid = player["id"]
    cur = int(player["peakRating"])

    if pid in MANUAL_LOCKS:
        return None, f"manual lock {MANUAL_LOCKS[pid]}"

    score = int((live or {}).get("score") or 0)
    last_rd = (live or {}).get("lastRd")

    # Named injury-aware corrections (restore / soften prior crush)
    if pid == "huddersfield-cur-matty-english":
        # Score 215, Last Rd 0 — wrongly cut 86→80 on season total
        return 85, "Injured/out (Last Rd 0, low Score); restore toward prior 86 with mild availability discount"

    if pid == "huddersfield-cur-niall-evalds":
        return 81, "Likely out / limited; restore prior 81 (was cut on season total)"

    if pid == "wigan-cur-liam-farrell":
        return 84, "Missed recent round / soft volume; soften prior cut (85->82) to 84"

    if pid == "hull-fc-cur-herman-eseese":
        # Playing (Last Rd > 0) but weak vs £150k — soften harsh 89→83
        return 85, "Available recently but weak vs price; soften prior cut (89->83) to 85"

    if pid == "wigan-cur-jai-field":
        if score >= 750 and (last_rd or 0) > 0:
            return 87, f"Updated Fantasy Score {score}, Last Rd {last_rd}; nudge 86->87"

    # Generic: if injured_out and we crushed far below pre-audit / xlsx OVR, restore
    if availability == "injured_out":
        xovr = None
        if xrow and xrow.get("OVR") is not None:
            try:
                xovr = int(xrow["OVR"])
            except (TypeError, ValueError):
                xovr = None
        restore_base = PRE_AUDIT_PEAK.get(pid) or xovr
        if restore_base is not None and cur <= restore_base - 3:
            tgt = max(70, min(95, restore_base - 1))
            if tgt != cur:
                return tgt, f"Injured/out restore toward {restore_base} (was crushed on season total)"
        return None, "Injured/out — leave current (already protected)"

    if availability == "limited":
        return None, "Limited — no further cut; keep current"

    return None, "Available — keep current calibrated rating"


def main() -> None:
    write_batch = "--write-batch" in __import__("sys").argv

    fantasy = parse_fantasy(FANTASY_MD.read_text(encoding="utf-8", errors="replace"))
    live_by = {f["norm"]: f for f in fantasy}
    xlsx = load_xlsx()
    current = load_current_players()

    # Snapshot for manual-preserve audit
    preserve = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "players": {
            p["id"]: {"name": p["name"], "peakRating": p["peakRating"], "club": p["club"]}
            for p in current
        },
    }
    OUT_PRESERVE.write_text(json.dumps(preserve, indent=2) + "\n", encoding="utf-8")

    prior_changes: dict[str, dict] = {}
    if PRIOR_REPORT.exists():
        prior = json.loads(PRIOR_REPORT.read_text(encoding="utf-8"))
        for c in prior.get("changes", []):
            prior_changes[c["id"]] = c

    changes: list[dict] = []
    preserved_manual: list[dict] = []
    classified: list[dict] = []
    skipped_manual_diverge: list[dict] = []

    for p in current:
        live = find_live(live_by, p["name"])
        xrow = xlsx.get(norm(p["name"]))
        availability, notes = classify(live, xrow)
        classified.append(
            {
                "id": p["id"],
                "name": p["name"],
                "peakRating": p["peakRating"],
                "availability": availability,
                "fantasyScore": (live or {}).get("score"),
                "lastRd": (live or {}).get("lastRd"),
                "metres": (live or {}).get("metres"),
                "price": (live or {}).get("priceRaw"),
                "notes": notes,
            }
        )

        if p["id"] in MANUAL_LOCKS:
            preserved_manual.append(
                {
                    "id": p["id"],
                    "name": p["name"],
                    "peakRating": p["peakRating"],
                    "reason": "explicit manual lock",
                }
            )
            # If somehow drifted from lock, force back to lock value
            lock = MANUAL_LOCKS[p["id"]]
            if p["peakRating"] != lock:
                changes.append(
                    {
                        "id": p["id"],
                        "name": p["name"],
                        "club": p["club"],
                        "from": p["peakRating"],
                        "to": lock,
                        "availability": availability,
                        "fantasyScore": (live or {}).get("score"),
                        "lastRd": (live or {}).get("lastRd"),
                        "notes": notes,
                        "reason": f"restore explicit manual lock {lock}",
                    }
                )
            continue

        tgt, reason = propose(p, live, xrow, availability)
        if tgt is None or tgt == p["peakRating"]:
            continue

        # If prior audit set a value and current differs, treat as manual and skip
        prior = prior_changes.get(p["id"])
        if prior and int(prior.get("to", p["peakRating"])) != p["peakRating"]:
            # Current already moved away from prior batch target — preserve
            skipped_manual_diverge.append(
                {
                    "id": p["id"],
                    "name": p["name"],
                    "current": p["peakRating"],
                    "priorBatchTo": prior.get("to"),
                    "proposed": tgt,
                    "reason": "current differs from prior batch target — preserve manual",
                }
            )
            continue

        changes.append(
            {
                "id": p["id"],
                "name": p["name"],
                "club": p["club"],
                "from": p["peakRating"],
                "to": tgt,
                "availability": availability,
                "fantasyScore": (live or {}).get("score"),
                "lastRd": (live or {}).get("lastRd"),
                "notes": notes,
                "reason": reason,
            }
        )

    changes.sort(key=lambda c: abs(c["to"] - c["from"]), reverse=True)

    report = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "methodology": {
            "fantasyM": "metres (not matches)",
            "availability": (
                "Inferred from Last Rd, Score vs Fantasy price band, metres volume, "
                "and xlsx injury/availability text"
            ),
            "rule": (
                "Do not crush injured/out players on season totals; "
                "restore toward pre-audit quality with mild availability discount"
            ),
            "manualLocks": MANUAL_LOCKS,
            "preserveSnapshot": str(OUT_PRESERVE.relative_to(ROOT)),
            "liveFantasy": str(FANTASY_MD),
        },
        "summary": {
            "currentPlayers": len(current),
            "liveFantasyRows": len(fantasy),
            "changeCount": len(changes),
            "injuredOutCount": sum(1 for c in classified if c["availability"] == "injured_out"),
            "limitedCount": sum(1 for c in classified if c["availability"] == "limited"),
            "preservedManualCount": len(preserved_manual),
            "skippedManualDivergeCount": len(skipped_manual_diverge),
        },
        "preservedManual": preserved_manual,
        "skippedManualDiverge": skipped_manual_diverge,
        "changes": changes,
        "priorFormCutsKept": [
            {
                "id": p["id"],
                "name": p["name"],
                "rating": p["peakRating"],
                "note": "Prior form cut kept — player available or intentional development rating",
            }
            for p in current
            if p["id"]
            in {
                "oldham-cur-jake-bibby",
                "castleford-cur-krystian-mapapalangi",
                "castleford-cur-jason-qareqare",
                "catalans-cur-benjamin-garcia",
                "leigh-cur-umyla-hanley",
                "bradford-cur-ethan-ryan",
                "bradford-cur-eribe-doro",
                "bradford-cur-jayden-nikorima",
                "castleford-cur-cain-robb",
                "castleford-cur-zac-cini",
                "catalans-cur-manase-kaho",
                "toulouse-cur-john-toleafoa",
                "wigan-cur-jonny-vaughan",
            }
        ],
        "injuredOutSample": [c for c in classified if c["availability"] == "injured_out"][:50],
        "limitedSample": [c for c in classified if c["availability"] == "limited"][:50],
    }
    OUT_REPORT.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_REPORT}")
    print(f"Changes: {len(changes)}")

    if write_batch:
        ratings = {c["id"]: c["to"] for c in changes}
        OUT_BATCH.write_text(
            json.dumps({"ratings": ratings, "potentials": {}}, indent=2) + "\n",
            encoding="utf-8",
        )
        print(f"Wrote {OUT_BATCH} ({len(ratings)} ratings)")

    for c in changes:
        safe_reason = c["reason"].encode("ascii", "replace").decode("ascii")
        print(
            f"  {c['name']} ({c['id']}): {c['from']} -> {c['to']} "
            f"[{c['availability']}] {safe_reason}"
        )


if __name__ == "__main__":
    main()
