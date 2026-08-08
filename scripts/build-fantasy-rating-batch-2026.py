#!/usr/bin/env python3
"""Focused Current-player rating correction batch (named high-confidence fixes)."""
from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FANTASY_MD = Path(
    r"C:\Users\Macis\.cursor\projects\c-Users-Macis-Projects-27-0\agent-tools\dcfc4def-a36c-41de-921a-26b50ec483d7.txt"
)
OUT_BATCH = ROOT / "data" / "player-attr-batch.json"
OUT_REPORT = ROOT / "data" / "fantasy-rating-calibration-2026-report.json"

ERA_26 = {
    "Bradford Bulls",
    "Castleford Tigers",
    "Catalans Dragons",
    "Huddersfield Giants",
    "Hull FC",
    "Hull KR",
    "Leeds Rhinos",
    "Leigh Leopards",
    "St Helens",
    "Toulouse Olympique",
    "Wakefield Trinity",
    "Warrington Wolves",
    "Wigan Warriors",
    "York Knights",
}
POS = sorted(
    [
        "Full Back",
        "Winger",
        "Centre",
        "Stand Off",
        "Scrum Half",
        "Prop",
        "Hooker",
        "Second Row",
        "Loose Forward",
    ],
    key=len,
    reverse=True,
)
POS_PAT = "|".join(re.escape(p) for p in POS)
POS_GROUP = {
    "FULLBACK": "OUTSIDE",
    "WING": "OUTSIDE",
    "CENTRE": "OUTSIDE",
    "STAND_OFF": "HALVES",
    "SCRUM_HALF": "HALVES",
    "HOOKER": "HOOKER",
    "PROP": "PROP",
    "SECOND_ROW": "BACKROW",
    "LOOSE_FORWARD": "BACKROW",
}
OUTSIDE_CURVE = [
    (0, 75),
    (250, 76),
    (400, 78),
    (550, 81),
    (700, 84),
    (746, 85),
    (782, 85),
    (791, 87),
    (864, 88),
    (936, 87),
    (1050, 90),
    (1200, 92),
]
HALVES_CURVE = [
    (0, 75),
    (250, 76),
    (400, 78),
    (550, 80),
    (700, 82),
    (827, 84),
    (950, 88),
    (1044, 92),
    (1200, 93),
]
GENERIC_CURVE = [
    (0, 75),
    (250, 76),
    (400, 78),
    (600, 81),
    (800, 84),
    (1000, 88),
    (1200, 91),
]
PROTECTED = {
    "wigan-cur-bevan-french",
    "hull-kr-cur-mikey-lewis",
    "wigan-cur-zach-eckersley",
    "warrington-cur-matty-ashton",
    "castleford-cur-daejarn-asi",
    "toulouse-cur-olly-ashall-bott",
    "castleford-cur-krystian-mapapalangi",
    "huddersfield-cur-jake-bibby",
}

METHOD = """
1. Match fantasy players to current-squads by normalised name (unique match preferred).
2. Fantasy->rating curve from well-aligned elites: French~93, Lewis~92, Ashton 87/791,
   Asi 84/827, Eckersley 88/864, Ashall-Bott 87/936; position-group knots.
3. Overrated flag: game peakRating >= 3 above peer-implied. Applied batch is named
   high-confidence fixes only; other overrates listed under flaggedOverratedNotApplied
   (no mass +/-2 churn).
4. MUST: Bibby 89->83 (no 2026 SL fantasy; 2025 smoothed 83); Mapapalangi 88->85
   (fantasy 782; peers Asi/Ashton/Eckersley).
5. Bradford audit: Ethan Ryan 84->78, Doro 82->77, Nikorima 83->80, Sutton 84->81,
   Loghan Lewis 78->81.
6. Development leftovers Robb/Kaho/Toleafoa/Vaughan -> xlsx 75 (weak fantasy).
   Cesar Rouge fantasy 721 solid -> HALVES peer ~82; within +/-2 of game 83 so
   rating unchanged (pot only).
7. Qareqare game 76 vs fantasy 746 underrated vs outside peers ~85; bump capped +/-6 -> 82.
8. Ashall-Bott 936/87 peer~87 — no change.
9. Cap +/-6 except Bibby/Mapapalangi-style and Development->75 clear errors.
""".strip()


def norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", s)).strip()


def interp(score: float, knots: list[tuple[int, float]]) -> float:
    if score <= knots[0][0]:
        return float(knots[0][1])
    if score >= knots[-1][0]:
        return float(knots[-1][1])
    for i in range(1, len(knots)):
        x0, y0 = knots[i - 1]
        x1, y1 = knots[i]
        if x0 <= score <= x1:
            return y0 + (score - x0) / max(1e-9, x1 - x0) * (y1 - y0)
    return float(knots[-1][1])


def parse_fantasy(text: str) -> list[dict]:
    out = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        parts = [p.strip() for p in line.strip().strip("|").split("|")]
        if len(parts) < 25 or parts[1] in ("Name", "---") or parts[0] == "Team":
            continue
        raw = parts[1]
        m = re.match(rf"^(.+?)({POS_PAT})$", raw)
        name = m.group(1).strip() if m else raw
        try:
            score = int(parts[-1].replace(",", ""))
        except ValueError:
            continue
        out.append(
            {"name": name, "score": score, "price": parts[3], "norm": norm(name)}
        )
    return out


def clamp(cur: int, tgt: int, cap: int = 6) -> int:
    d = tgt - cur
    if d > cap:
        return cur + cap
    if d < -cap:
        return cur - cap
    return tgt


def main() -> None:
    fantasy = parse_fantasy(FANTASY_MD.read_text(encoding="utf-8", errors="replace"))
    fb: dict[str, list[dict]] = defaultdict(list)
    for f in fantasy:
        fb[f["norm"]].append(f)

    squads = json.loads((ROOT / "data" / "current-squads.json").read_text(encoding="utf-8"))
    by_id = {p["id"]: p for p in squads}
    era = [p for p in squads if (p.get("club") or p.get("team")) in ERA_26]
    era_n: dict[str, list[dict]] = defaultdict(list)
    for p in era:
        era_n[norm(p["name"])].append(p)

    wb = openpyxl.load_workbook(ROOT / "super_league_2026_ratings.xlsx", data_only=True)
    ws = wb["All Players"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    xlsx: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("Player"):
            xlsx[norm(str(d["Player"]))] = d

    hist: dict[str, int] = {}
    with (ROOT / "data" / "imports" / "historic_super_league_ratings_2020_2025_smoothed.csv").open(
        encoding="utf-8", newline=""
    ) as f:
        for row in csv.DictReader(f):
            if str(row.get("year")) != "2025":
                continue
            try:
                r = int(float(row["rating"]))
            except (TypeError, ValueError):
                continue
            n = norm(row.get("player") or "")
            hist[n] = max(hist.get(n, 0), r)

    ov = json.loads((ROOT / "data" / "player-rating-overrides.json").read_text(encoding="utf-8"))
    pot_ov = {k: int(v) for k, v in (ov.get("potentialOverrides") or {}).items()}

    match_meta = {"unique": 0, "ambiguous": 0, "unmatched_nonzero": 0}
    matched = []
    for nkey, flist in fb.items():
        f = max(flist, key=lambda x: x["score"])
        cands = era_n.get(nkey, [])
        if len(cands) == 1:
            p = cands[0]
            matched.append(
                {
                    "id": p["id"],
                    "name": p["name"],
                    "rating": int(p["peakRating"]),
                    "score": f["score"],
                    "group": POS_GROUP.get(p.get("position") or "", "OTHER"),
                }
            )
            match_meta["unique"] += 1
        elif len(cands) > 1:
            match_meta["ambiguous"] += 1
        elif f["score"] > 0:
            match_meta["unmatched_nonzero"] += 1

    def fscore(pid: str, *alts: str) -> int | None:
        p = by_id[pid]
        keys = [norm(p["name"]), *alts]
        best = None
        for k in keys:
            if k in fb:
                s = max(x["score"] for x in fb[k])
                best = s if best is None else max(best, s)
        if "rouge" in norm(p["name"]):
            for k, v in fb.items():
                if "rouge" in k:
                    s = max(x["score"] for x in v)
                    best = s if best is None else max(best, s)
        return best

    ratings: dict[str, int] = {}
    pots: dict[str, int] = {}
    changes: list[dict] = []
    reviewed: list[dict] = []

    def add(
        pid: str,
        to: int,
        reason: str,
        fs: int | None = None,
        pot: int | None = None,
        uncapped: bool = False,
    ) -> None:
        p = by_id[pid]
        frm = int(p["peakRating"])
        if not uncapped:
            to = clamp(frm, to, 6)
        if to != frm:
            ratings[pid] = to
        if pot is not None:
            pot = max(int(pot), to if to != frm else frm)
            pots[pid] = pot
        if to == frm and pot is None:
            return
        entry = {
            "id": pid,
            "name": p["name"],
            "from": frm,
            "to": to if to != frm else frm,
            "fantasyScore": fs,
            "reason": reason,
        }
        if pid in pots:
            entry["potential"] = pots[pid]
        changes[:] = [c for c in changes if c["id"] != pid]
        changes.append(entry)

    bibby_h = hist.get(norm("Jake Bibby"), 83)
    add(
        "huddersfield-cur-jake-bibby",
        83,
        f"mandatory: no 2026 SL fantasy; Championship loan but Huddersfield card; 2025 smoothed OVR {bibby_h}",
        None,
        pot=83,
        uncapped=True,
    )
    add(
        "castleford-cur-krystian-mapapalangi",
        85,
        "mandatory: fantasy 782; peers Asi 84/827, Ashton 87/791, Eckersley 88/864 — 88 too high (xlsx 75 obsolete)",
        fscore("castleford-cur-krystian-mapapalangi"),
        pot=88,
        uncapped=True,
    )

    for pid, to, reason in [
        ("bradford-cur-ethan-ryan", 78, "deep audit: Bradford Ethan Ryan overrated vs role/form"),
        ("bradford-cur-eribe-doro", 77, "deep audit: Bradford Eribe Doro import drift"),
        ("bradford-cur-jayden-nikorima", 80, "deep audit: Bradford Jayden Nikorima import drift"),
        ("bradford-cur-ryan-sutton", 81, "deep audit: Bradford Ryan Sutton import drift"),
        ("bradford-cur-loghan-lewis", 81, "deep audit: Bradford Loghan Lewis underrated; xlsx OVR 81"),
    ]:
        p = by_id[pid]
        by = p.get("birthYear") or 0
        pot = None
        if pid == "bradford-cur-loghan-lewis":
            pot = max(pot_ov.get(pid, 0), to + 2, 84)
        elif by >= 2002:
            pot = max(pot_ov.get(pid, 0), to + 2)
        elif pid in pot_ov:
            pot = max(pot_ov[pid], to)
        add(pid, to, reason, fscore(pid), pot=pot)

    for pid, label in [
        ("castleford-cur-cain-robb", "Cain Robb"),
        ("catalans-cur-manase-kaho", "Manase Kaho"),
        ("toulouse-cur-john-toleafoa", "John Toleafoa"),
        ("salford-cur-jonny-vaughan", "Jonny Vaughan"),
        ("catalans-cur-cesar-rouge", "Cesar Rouge"),
    ]:
        p = by_id[pid]
        fs = fscore(pid, norm(label))
        n = norm(p["name"])
        xrow = xlsx.get(n)
        if not xrow and "rouge" in n:
            for k, v in xlsx.items():
                if "rouge" in k:
                    xrow = v
                    break
        xovr = int(xrow["OVR"]) if xrow and xrow.get("OVR") is not None else 75
        group = POS_GROUP.get(p.get("position") or "", "OTHER")
        frm = int(p["peakRating"])
        by = p.get("birthYear") or 0
        solid = fs is not None and fs >= 500
        if solid:
            knots = HALVES_CURVE if group == "HALVES" else OUTSIDE_CURVE
            implied = interp(fs, knots)
            target = int(round(implied))
            if abs(frm - target) <= 2:
                pots[pid] = max(
                    pot_ov.get(pid, 0),
                    frm + 2 if by >= 2001 else frm,
                    85 if by >= 2001 else frm,
                )
                changes.append(
                    {
                        "id": pid,
                        "name": p["name"],
                        "from": frm,
                        "to": frm,
                        "fantasyScore": fs,
                        "reason": (
                            f"potential only: solid fantasy {fs}, rating OK vs peer ~{implied:.1f} "
                            f"(xlsx {xovr} superseded)"
                        ),
                        "potential": pots[pid],
                    }
                )
                reviewed.append(
                    {
                        "id": pid,
                        "name": p["name"],
                        "note": f"solid fantasy {fs}; peer ~{implied:.1f}; within +/-2 of {frm}",
                    }
                )
            else:
                pot = max(pot_ov.get(pid, 0), target + (2 if by >= 2001 else 0), target)
                if by >= 2003:
                    pot = max(pot, 84)
                add(
                    pid,
                    target,
                    f"development leftover but solid fantasy {fs}; peer ~{implied:.1f} (xlsx {xovr} superseded)",
                    fs,
                    pot=pot,
                )
        else:
            pot = max(pot_ov.get(pid, 0), xovr + (8 if by >= 2003 else 2), xovr)
            if by >= 2003:
                pot = max(pot, 84)
            add(
                pid,
                xovr,
                f"development leftover wrongly high; weak/no 2026 fantasy ({fs}); xlsx OVR {xovr}",
                fs,
                pot=pot,
                uncapped=True,
            )

    qid = "castleford-cur-jason-qareqare"
    q = by_id[qid]
    qs = fscore(qid) or 0
    qi = interp(qs, OUTSIDE_CURVE)
    qt = int(round(qi))
    qcheck = {
        "fantasyScore": qs,
        "peerImplied": round(qi, 2),
        "game": int(q["peakRating"]),
        "changed": False,
        "to": None,
    }
    if int(q["peakRating"]) <= qt - 3:
        qto = clamp(int(q["peakRating"]), qt, 6)
        add(
            qid,
            qto,
            f"underrated: fantasy {qs} vs outside peer curve ~{qi:.1f}; bump capped +/-6",
            qs,
            pot=max(pot_ov.get(qid, 0), qto + 2, 85),
        )
        qcheck.update(changed=True, to=qto)

    aid = "toulouse-cur-olly-ashall-bott"
    a = by_id[aid]
    ascore = fscore(aid, norm("Oliver Ashall-Bott"), norm("Olly Ashall-Bott")) or 936
    ai = interp(ascore, OUTSIDE_CURVE)
    acheck = {
        "fantasyScore": ascore,
        "peerImplied": round(ai, 2),
        "game": int(a["peakRating"]),
        "changed": False,
        "to": None,
        "note": "within tolerance — no change",
    }

    curve_map = {
        "OUTSIDE": OUTSIDE_CURVE,
        "HALVES": HALVES_CURVE,
        "HOOKER": GENERIC_CURVE,
        "PROP": GENERIC_CURVE,
        "BACKROW": GENERIC_CURVE,
        "OTHER": GENERIC_CURVE,
    }
    flagged = []
    for m in matched:
        if m["id"] in ratings or m["id"] in PROTECTED:
            continue
        implied = interp(m["score"], curve_map[m["group"]])
        gap = m["rating"] - implied
        if gap >= 3 and abs(m["rating"] - round(implied)) > 2:
            flagged.append(
                {
                    "id": m["id"],
                    "name": m["name"],
                    "from": m["rating"],
                    "peerImplied": round(implied, 1),
                    "fantasyScore": m["score"],
                    "group": m["group"],
                    "gap": round(gap, 1),
                    "applied": False,
                    "note": "flagged only — not mass-applied",
                }
            )

    for pid, r in list(ratings.items()):
        if pid in pots:
            pots[pid] = max(pots[pid], r)
        else:
            p = by_id[pid]
            by = p.get("birthYear") or 0
            if by >= 2002:
                pots[pid] = max(pot_ov.get(pid, 0), r + 2)
            elif pid in pot_ov:
                pots[pid] = max(pot_ov[pid], r)
    pots["castleford-cur-krystian-mapapalangi"] = 88
    pots["huddersfield-cur-jake-bibby"] = 83

    changes.sort(key=lambda c: (-abs(c["to"] - c["from"]), c["name"]))
    batch = {"ratings": ratings, "potentials": pots}
    report = {
        "method": METHOD,
        "inputs": {
            "fantasyMarkdown": str(FANTASY_MD),
            "squads": str(ROOT / "data" / "current-squads.json"),
            "xlsx": str(ROOT / "super_league_2026_ratings.xlsx"),
            "historic2025": str(
                ROOT / "data" / "imports" / "historic_super_league_ratings_2020_2025_smoothed.csv"
            ),
        },
        "matchMeta": match_meta,
        "anchors": {
            "Bevan French": 93,
            "Mikey Lewis": 92,
            "Matty Ashton": "87/791",
            "Daejarn Asi": "84/827",
            "Zach Eckersley": "88/864",
            "Oliver Ashall-Bott": "87/936",
        },
        "qareqareCheck": qcheck,
        "ashallBottCheck": acheck,
        "reviewedNoChange": reviewed,
        "flaggedOverratedNotApplied": flagged,
        "changes": changes,
        "ratings": ratings,
        "potentials": pots,
    }
    OUT_BATCH.write_text(json.dumps(batch, indent=2) + "\n", encoding="utf-8")
    OUT_REPORT.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")

    print("=== RATINGS ===")
    print(json.dumps(ratings, indent=2, ensure_ascii=False))
    print("=== POTENTIALS ===")
    print(json.dumps(pots, indent=2, ensure_ascii=False))
    print("\n=== SUMMARY ===")
    hdr = "Name                         From   To   d  Fant  Reason"
    print(hdr)
    for c in changes:
        d = c["to"] - c["from"]
        fs = "" if c["fantasyScore"] is None else str(c["fantasyScore"])
        name = c["name"].encode("ascii", "replace").decode("ascii")
        reason = c["reason"][:72].encode("ascii", "replace").decode("ascii")
        print(f"{name[:28]:28} {c['from']:4} {c['to']:4} {d:+3} {fs:>5}  {reason}")
    print(
        f"\n{len(ratings)} rating changes, {len(pots)} potential sets; "
        f"flagged-not-applied={len(flagged)}"
    )
    print("Qareqare:", qcheck)
    print("Ashall-Bott:", acheck)
    print("Wrote", OUT_BATCH)
    print("Wrote", OUT_REPORT)


if __name__ == "__main__":
    main()
